# This script takes the list of specimens to be re-imaged from the QA_Images_Issues.xlsx file and searches for the corresponding barcode in the DigiApp exports.
# It also searches for missing barcodes in the SQLite databases for each workstation.
# It extracts the taxonfullname, storagefullname, and storagename columns from the DigiApp exports and writes them to a new sheet in the same Excel file.

import pandas as pd
import os
import glob
import sqlite3
from openpyxl import load_workbook
from datetime import datetime
from dotenv import load_dotenv

# Load environment variables from the .env file
load_dotenv()

# file_path = 'N:/SCI-SNM-DigitalCollections/DaSSCo/Workflows and workstations/Quality assurance and control/QA_Images_Issues.xlsx'  
file_path = os.getenv("FILE_PATH")
base_directory = os.getenv("BASE_DIRECTORY")  
db_directory = os.getenv("DB_DIRECTORY")

def get_collection(workstation):
    mapping = {
        'WORKHERB0001': 'NHMD_Herbarium',
        'WORKHERB0002': 'AU_Herbarium',
        'WORKHERB0003': 'NHMD_Herbarium',
        'WORKPIOF0001': 'NHMD_PinnedInsects',
        'WORKPIOF0002': 'NHMD_PinnedInsects',
        'WORKPIOF0003': 'NHMA_PinnedInsects'
    }
    return mapping.get(workstation, None)

# Pull the taxonomy and storage information from the DigiApp exports based on the barcode
def search_barcode_in_csv(barcode, directory):
    if not barcode or barcode.upper() == 'NA':
        return None
    
    try:
        # csv_files = glob.glob(os.path.join(directory, '*_original.csv'))
        barcode_trimmed = str(barcode).lstrip('0')  # Ensure barcode is a string and remove leading zeros
        
        # Walk through all subdirectories in 'directory'
        for dirpath, dirnames, filenames in os.walk(directory):
            csv_files = glob.glob(os.path.join(dirpath, '*_original.csv'))

            for file in csv_files:
                try:
                    digiapp_exports_df = pd.read_csv(file, delimiter=';', dtype=str)
                    if 'catalognumber' in digiapp_exports_df.columns:
                        digiapp_exports_df['catalognumber_trimmed'] = digiapp_exports_df['catalognumber'].astype(str).str.lstrip('0')
                        if barcode_trimmed in digiapp_exports_df['catalognumber_trimmed'].values:
                            matching_row = digiapp_exports_df[digiapp_exports_df['catalognumber_trimmed'] == barcode_trimmed]
                            return {
                                'filename': file,
                                'taxonfullname': matching_row['taxonfullname'].values[0] if 'taxonfullname' in matching_row.columns else None,
                                'storagefullname': matching_row['storagefullname'].values[0] if 'storagefullname' in matching_row.columns else None,
                                'storagename': matching_row['storagename'].values[0] if 'storagename' in matching_row.columns else None
                            }
                except Exception as e:
                    print(f"Error reading {file}: {e}")
    except Exception as e:
        print(f"Error searching for barcode {barcode}: {e}")
    
    return None

# Pull barcode and date_asset_taken from the SQLite database based on the GUID
def query_barcodes_and_dates_from_db(workstation, guid, db_directory):
    if not guid or pd.isna(guid) or guid.strip() == "":
        print(f"Skipping empty GUID for workstation {workstation}")
        return [], []

    db_path = os.path.join(db_directory, f"{workstation}_jpgs.db")
    if not os.path.exists(db_path):
        print(f"Database not found: {db_path}")
        return [], []

    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        query = "SELECT barcode, date_asset_taken FROM table1 WHERE GUID = ?"
        cursor.execute(query, (guid,))
        results = cursor.fetchall()
        conn.close()

        barcodes = []
        dates = []
        for result in results:
            barcode = result[0]
            date_taken = result[1]
            barcode = barcode.strip("[]").strip("'\"")
            barcode = barcode.lstrip('0')
            if barcode:
                barcodes.append(barcode)
            if date_taken:
                dates.append(date_taken)

        print(f"Found barcodes {barcodes} and dates {dates} for GUID {guid} in {db_path}")
        return barcodes, dates
    except Exception as e:
        print(f"Error querying database {db_path} with GUID {guid}: {e}")
        return [], []

# Pull the list of specimens to be re-imaged from the QA_Images_Issues file based on the 'Follow-up Action Required' column
def read_specimens_xlsx(file_path, base_directory, db_directory):
    try:
        xls = pd.ExcelFile(file_path)
        df = pd.read_excel(
            xls,
            sheet_name='Specimens',
            usecols=[
                'Workstation', 'Follow-up Action Required', 'GUID',
                'Folder Date: Year', 'Folder Date: Month', 'Folder Date: Day', 'Barcode'  # <-- include Barcode if it exists
            ],
            dtype={'GUID': str, 'Barcode': str}
        )

        df = df[df['Follow-up Action Required'].str.contains('re-image', case=False, na=False)]

        # --- Only query DB if Barcode is null or empty ---
        def get_barcodes_and_dates(row):
            barcode_val = str(row.get('Barcode', '')).strip()

            # Skip querying DB if barcode already present
            if barcode_val and barcode_val.upper() not in ['NA', 'NAN']:
                print(f"Skipping DB query — barcode already present for GUID {row['GUID']}")
                return pd.Series([[], []])

            # Query DB only when Barcode is empty
            return pd.Series(query_barcodes_and_dates_from_db(row['Workstation'], row['GUID'], db_directory))

        df[['Barcodes_from_DB', 'Dates_from_DB']] = df.apply(get_barcodes_and_dates, axis=1)
        # Keep rows that have either a Barcode already or DB results
        df = df[df['Barcodes_from_DB'].apply(lambda x: len(x) > 0) | df['Barcode'].notna()]

        # Merge DB results into Barcode/Date columns
        df['Barcode'] = df.apply(
            lambda row: ';'.join(row['Barcodes_from_DB']) if isinstance(row['Barcodes_from_DB'], list) and len(row['Barcodes_from_DB']) > 0 else row.get('Barcode', ''),
            axis=1
        )
        df['Date_Asset_Taken'] = df['Dates_from_DB'].apply(lambda x: ';'.join(x) if isinstance(x, list) else str(x))

        def search_and_extract(row):
            collection = get_collection(row['Workstation'])
            # --- Only run search_barcode_in_csv if collection != AU_Herbarium ---
            if collection == 'AU_Herbarium':
                print(f"Skipping DigiApp search for AU_Herbarium (Workstation: {row['Workstation']})")
                return pd.Series({
                    'filename': None,
                    'taxonfullname': None,
                    'storagefullname': None,
                    'storagename': None
                })

    
            barcode_list = []
            if isinstance(row.get('Barcodes_from_DB'), list) and len(row['Barcodes_from_DB']) > 0:
                barcode_list = row['Barcodes_from_DB']
            elif row.get('Barcode'):
                barcode_list = [row['Barcode']]

            if not barcode_list:
                return pd.Series({'filename': None, 'taxonfullname': None, 'storagefullname': None, 'storagename': None})

            if collection:
                directory = os.path.join(base_directory, '6.Archive', collection)
                result = search_barcode_in_csv(barcode_list[0], directory)
                if result:
                    return pd.Series(result)

            return pd.Series({'filename': None, 'taxonfullname': None, 'storagefullname': None, 'storagename': None})

        extracted_data = df.apply(search_and_extract, axis=1)
        new_df = pd.concat([df, extracted_data], axis=1)

        # Ensure column order: reimaged (blank), barcode, taxonfullname, storagefullname, then all others
        new_df['reimaged'] = ''  # create new blank column for tracking status if not present
        new_df['date_reimaged'] = ''  # create new blank column for tracking date_reimaged if not present

        # Define preferred column order
        preferred_order = ['reimaged', 'date_reimaged', 'Barcode', 'taxonfullname', 'storagename', 'Follow-up Action Required']

        # Keep preferred columns first (if they exist), then all remaining columns
        ordered_cols = [col for col in preferred_order if col in new_df.columns] + \
                    [col for col in new_df.columns if col not in preferred_order]

        # Reorder DataFrame
        new_df = new_df[ordered_cols]

        # --- Write results to Excel ---
        collections = df['Workstation'].apply(get_collection).unique()
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            for collection in collections:
                collection_df = new_df[df['Workstation'].apply(get_collection) == collection]
                sheet_name = f'Reimage_Needed_{collection}'
                try:
                    existing_df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=str)
                    collection_df = collection_df[~collection_df['GUID'].isin(existing_df['GUID'])]
                    combined_df = pd.concat([existing_df, collection_df], ignore_index=True)
                except ValueError:
                    combined_df = collection_df

                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)

        return new_df

    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return pd.DataFrame()
    
# Add excel formulas
def add_excel_formulas(file_path):
    wb = load_workbook(file_path)

    # --- Add date_reimaged formulas ---
    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("Reimage_Needed_"):
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            if 'reimaged' in headers and 'date_reimaged' in headers:
                reimaged_col = headers.index('reimaged') + 1
                date_col = headers.index('date_reimaged') + 1
                for row in range(2, ws.max_row + 1):
                    reimaged_coord = ws.cell(row=row, column=reimaged_col).coordinate
                    # Use English formulas with commas
                    ws.cell(row=row, column=date_col).value = (
                        f'=IF(LOWER({reimaged_coord})="yes",TODAY(),"")'
                    )
                    ws.cell(row=row, column=date_col).number_format = "YYYY-MM-DD"

    wb.save(file_path)


df = read_specimens_xlsx(file_path, base_directory, db_directory)
print(df.head())
add_excel_formulas(file_path)